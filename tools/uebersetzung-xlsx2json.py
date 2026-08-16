#!/usr/bin/env python3
"""Konvertiert die ausgefüllte Übersetzungsvorlage (XLSX) nach translations-en.json."""
import json, sys
from openpyxl import load_workbook

src, dst = sys.argv[1], sys.argv[2]
wb = load_workbook(src, data_only=True)
rows = []
def val(c): return str(c or '').strip()
for sheet, ref_col, en_col in [('Feste Seiten', 5, 4), ('Einzelleistungen', 5, 4), ('Team', 5, 4), ('Stellenangebote', 5, 4), ('Karriere & Footer', 5, 4)]:
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        en = val(ws.cell(r, en_col).value)
        if not en or 'BEISPIEL' in en: continue
        rows.append({'sheet': sheet, 'ref': val(ws.cell(r, ref_col).value), 'en': en})
ws = wb['URL-Kürzel']
for r in range(2, ws.max_row + 1):
    en = val(ws.cell(r, 3).value).split('←')[0].strip()
    if not en: continue
    rows.append({'sheet': 'URL-Kürzel', 'typ': val(ws.cell(r, 1).value), 'de': val(ws.cell(r, 2).value), 'en': en})
json.dump(rows, open(dst, 'w'), ensure_ascii=False, indent=1)
print(len(rows), 'Übersetzungen ->', dst)
